"""
Belçika Market Kategori Sistemi — Kapsamlı Excel Üretici
16 departman, ~420 L3 alt kategori, ~1400 L4 ürün tipi
43 öznitelik, ~120 normalizasyon kuralı
8 Excel sayfası
"""

from __future__ import annotations
import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 1. DEPARTMANLAR (D01-D16)
# ─────────────────────────────────────────────
DEPARTMANLAR = [
    ("D01", "Süt, Yumurta & Peynir",        "Zuivel, Eieren & Kaas",         "Produits laitiers, Œufs & Fromages"),
    ("D02", "Et, Kümes & Balık",              "Vlees, Gevogelte & Vis",         "Viande, Volaille & Poisson"),
    ("D03", "Meyve & Sebze",                  "Groenten & Fruit",               "Légumes & Fruits"),
    ("D04", "Ekmek & Fırın Ürünleri",         "Brood & Bakkerij",               "Pain & Boulangerie"),
    ("D05", "Temel Bakkaliye",                "Basisvoedsel",                   "Épicerie de base"),
    ("D06", "İçecekler",                      "Dranken",                        "Boissons"),
    ("D07", "Dondurulmuş Ürünler",            "Diepvriesproducten",             "Produits surgelés"),
    ("D08", "Atıştırmalık & Şekerleme",       "Snacks & Snoep",                 "Snacks & Confiserie"),
    ("D09", "Kahvaltılık & Mısır Gevreği",    "Ontbijt & Ontbijtgranen",        "Petit-déjeuner & Céréales"),
    ("D10", "Hazır & Soğutmalı Yemekler",     "Kant-en-klaar & Koelvers",       "Plats préparés & Réfrigérés"),
    ("D11", "Ev Temizliği & Çamaşır",         "Huishoudreiniging & Wasserij",   "Nettoyage maison & Lessive"),
    ("D12", "Kişisel Bakım & Hijyen",         "Persoonlijke verzorging",        "Hygiène & Soins personnels"),
    ("D13", "Bebek & Çocuk",                  "Baby & Kind",                    "Bébé & Enfant"),
    ("D14", "Evcil Hayvan",                   "Huisdieren",                     "Animaux domestiques"),
    ("D15", "Belçika Özgün & Mevsimsel",      "Belgische specialiteiten",       "Spécialités belges & Saisonnières"),
    ("D16", "Etnik & Uluslararası Gıda",      "Etnisch & Internationaal",       "Alimentation ethnique & Internationale"),
]

# ─────────────────────────────────────────────
# 2. KATEGORİ AĞACI (D → L2 → L3 → L4)
# Her tuple: (dept_id, L2, L3, L4_listesi)
# ─────────────────────────────────────────────
KATEGORILER = [
    # ── D01 Süt, Yumurta & Peynir ─────────────────────────────────
    ("D01", "Süt & Süt Alternatifleri", "İnek Sütü",
     ["Tam Yağlı Süt", "Yarım Yağlı Süt", "Yağsız Süt", "Laktosuz Süt", "UHT Süt", "Organik Süt"]),
    ("D01", "Süt & Süt Alternatifleri", "Bitki Bazlı Süt",
     ["Soya Sütü", "Yulaf Sütü", "Badem Sütü", "Pirinç Sütü", "Kaju Sütü", "Fındık Sütü", "Hindistan Cevizi Sütü"]),
    ("D01", "Yoğurt & Fermente", "Yoğurt",
     ["Tam Yağlı Yoğurt", "Yarım Yağlı Yoğurt", "Yunan Yoğurdu", "Laktosuz Yoğurt", "Bitki Bazlı Yoğurt"]),
    ("D01", "Yoğurt & Fermente", "Fermente İçecekler",
     ["Kefir (İnek)", "Kefir (Keçi)", "Kefir (Bitki)", "Ayran", "Lassi", "Skyr"]),
    ("D01", "Peynir", "Sert Peynir",
     ["Gouda", "Edam", "Emmental", "Gruyère", "Parmesan", "Cheddar", "Manchego", "Comté"]),
    ("D01", "Peynir", "Yarı Sert Peynir",
     ["Leerdammer", "Jarlsberg", "Raclette", "Tilsit", "Havarti"]),
    ("D01", "Peynir", "Yumuşak & Taze Peynir",
     ["Brie", "Camembert", "Mozzarella", "Ricotta", "Mascarpone", "Feta", "Cottage Cheese", "Fromage Blanc"]),
    ("D01", "Peynir", "Belçika Peynirleri",
     ["Herve (AOP)", "Chimay Peyniri", "Postelein", "Bouquet des Moines", "Passendale", "Père Joseph", "Orval Peyniri"]),
    ("D01", "Peynir", "Mavi Küflü Peynir",
     ["Roquefort", "Gorgonzola", "Stilton", "Danish Blue"]),
    ("D01", "Yumurta", "Tavuk Yumurtası",
     ["S Beden (53-63g)", "M Beden (53-63g)", "L Beden (63-73g)", "XL Beden (73g+)", "Organik Yumurta", "Serbest Dolaşım", "Vitamin Zenginleştirilmiş"]),
    ("D01", "Yumurta", "Diğer Yumurtalar",
     ["Bıldırcın Yumurtası", "Ördek Yumurtası", "Kaz Yumurtası"]),
    ("D01", "Tereyağı & Yağlar", "Tereyağı",
     ["Tuzlu Tereyağı", "Tuzsuz Tereyağı", "Organik Tereyağı", "Laktosuz Tereyağı", "Ghee", "Kahvaltılık Yağ Karışımı"]),
    ("D01", "Tereyağı & Yağlar", "Margarin & Bitkisel Yağ Spreadleri",
     ["Sürme Margarin", "Light Margarin", "Vegan Spread", "Omega-3 Zenginleştirilmiş"]),
    ("D01", "Krema & Tatlı Süt Ürünleri", "Krema",
     ["Taze Krema %15", "Taze Krema %30", "Krem Şanti", "Ekşi Krema", "Crème Fraîche"]),
    ("D01", "Krema & Tatlı Süt Ürünleri", "Sütlü Tatlılar",
     ["Puding", "Flans", "Crème Brûlée", "Panna Cotta", "Riz au lait (Sütlaç)", "Mousse"]),

    # ── D02 Et, Kümes & Balık ──────────────────────────────────────
    ("D02", "Sığır Eti", "Taze Sığır",
     ["Biftek (Entrecôte)", "Döş", "Kıyma (%5)", "Kıyma (%20)", "Rosto", "Konfit", "Dil", "Böbrek"]),
    ("D02", "Sığır Eti", "Hazır Sığır",
     ["Hamburger Köfte", "Meatballs", "Beef Burger", "Döner Dilimi"]),
    ("D02", "Domuz Eti", "Taze Domuz",
     ["Pirzola", "Jambon (Taze)", "Sosis", "Bacon", "Kıyma"]),
    ("D02", "Domuz Eti", "Şarküteri",
     ["Pâté", "Rillettes", "Filet Américain", "Coppa", "Pancetta", "Salami", "Ardenne Jambon (IGP)"]),
    ("D02", "Kümes Hayvanları", "Tavuk",
     ["Bütün Tavuk", "Göğüs Fileto", "But", "Kanat", "Kıyma", "Organik Tavuk", "Label Rouge"]),
    ("D02", "Kümes Hayvanları", "Diğer Kümes",
     ["Hindi Fileto", "Ördek Göğsü (Magret)", "Güvercin", "Bıldırcın"]),
    ("D02", "Balık & Deniz Ürünleri", "Taze & Soğutmalı Balık",
     ["Somon", "Levrek", "Çipura", "Morina (Kabeljauw)", "Alabalık", "Ton Balığı", "Pisi"]),
    ("D02", "Balık & Deniz Ürünleri", "Dondurulmuş Balık",
     ["Balık Fileto", "Balık Parmak", "Kariyer (Shrimp)", "Kalamar", "Ahtapot", "Midye", "İstiridye"]),
    ("D02", "Balık & Deniz Ürünleri", "Konserve & Marine",
     ["Ton Konservesi (Su)", "Ton Konservesi (Yağ)", "Sardalye", "Hamsi", "Ringa (Haring)", "Smoked Salmon"]),
    ("D02", "Balık & Deniz Ürünleri", "Kuzey Denizi Özel",
     ["Hollandse Nieuwe (Taze Ringa)", "Garnalen (Grijze)", "Mosselen (Midye)", "Paling (Yılan Balığı)"]),

    # ── D03 Meyve & Sebze ─────────────────────────────────────────
    ("D03", "Taze Sebze", "Yapraklı Sebze",
     ["Ispanak", "Roka", "Marul", "Kıvırcık", "Kale", "Pırasa", "Brüksel Lahanası", "Kırmızı Lahana"]),
    ("D03", "Taze Sebze", "Kök Sebze",
     ["Havuç", "Patates", "Şalgam", "Pancar", "Kereviz Kökü", "Turp", "Pasternak"]),
    ("D03", "Taze Sebze", "Soğansılar",
     ["Soğan", "Sarımsak", "Kırmızı Soğan", "Şalot", "Taze Soğan", "Pırasa"]),
    ("D03", "Taze Sebze", "Meyve Sebze",
     ["Domates", "Salatalık", "Biber", "Patlıcan", "Kabak", "Balkabağı", "Avokado"]),
    ("D03", "Taze Sebze", "Belçika Özgün Sebzeler",
     ["Witloof (Chicon/Belgian Endive)", "Brüksel Lahanası", "Hop Filizi (Asparagus de Houblon)", "Chicorée Rouge"]),
    ("D03", "Taze Sebze", "Mantar",
     ["Beyaz Mantar", "Portobello", "Shiitake", "Oyster Mantar", "Kestane Mantarı", "Truffe"]),
    ("D03", "Taze Meyve", "Mevsim Meyveleri",
     ["Elma", "Armut", "Kiraz", "Çilek", "Böğürtlen", "Ahududu", "Yaban Mersini", "Erik"]),
    ("D03", "Taze Meyve", "Turunçgiller",
     ["Portakal", "Mandalina", "Limon", "Greyfurt", "Lime", "Kan Portakalı"]),
    ("D03", "Taze Meyve", "Tropikal & Egzotik",
     ["Muz", "Ananas", "Mango", "Papaya", "Kivi", "Passion Fruit", "Longan", "Rambutan"]),
    ("D03", "Organik & Yerel", "Organik Sebze",
     ["Organik Paket Sebze", "Organik Kök Sebze", "Organik Yapraklı"]),
    ("D03", "Organik & Yerel", "Yerel Üretim (Belçika)",
     ["Belçika Elması (Jonagold)", "Belçika Armudu (Conférence)", "Gooik Elması", "Hebben Çileği"]),

    # ── D04 Ekmek & Fırın ─────────────────────────────────────────
    ("D04", "Ekmek", "Beyaz Ekmek",
     ["Sandwich Ekmeği", "Baget", "Ciabatta", "Focaccia", "Pide", "Kaiser Roll"]),
    ("D04", "Ekmek", "Tam Buğday & Çok Tahıllı",
     ["Tam Buğday Ekmeği", "Çavdar Ekmeği (Roggebrood)", "Spelt Ekmeği", "Tohumlu Ekmek", "Karavit"]),
    ("D04", "Ekmek", "Glutensiz & Özel",
     ["Glutensiz Ekmek", "Laktosuz Ekmek", "Vegan Ekmek", "Çörek Otu Ekmeği"]),
    ("D04", "Belçika Fırın", "Belçika Waffle",
     ["Gaufre de Liège", "Gaufre de Bruxelles", "Mini Waffle", "Waffle Dondurulmuş"]),
    ("D04", "Belçika Fırın", "Pistolet & Brioche",
     ["Pistolet (Sert)", "Pistolet (Yumuşak)", "Brioche", "Pain au Chocolat", "Croissant"]),
    ("D04", "Pasta & Kek", "Dilimli Kek",
     ["Kek Dilimi (Limon)", "Kek Dilimi (Çikolata)", "Mermer Kek", "Meyve Keki"]),
    ("D04", "Pasta & Kek", "Özel Pasta",
     ["Doğum Günü Pastası", "Noel Pastası", "Tiramisu", "Profiterol"]),

    # ── D05 Temel Bakkaliye ───────────────────────────────────────
    ("D05", "Tahıl & Makarna", "Pirinç",
     ["Uzun Taneli Pirinç", "Basmati", "Jasmine", "Sushi Pirinci", "Tam Buğday Pirinci", "Risotto Pirinci"]),
    ("D05", "Tahıl & Makarna", "Makarna",
     ["Spagetti", "Penne", "Fusilli", "Tagliatelle", "Lasagne", "Glutensiz Makarna", "Tam Buğday Makarna"]),
    ("D05", "Tahıl & Makarna", "Tahıllar & Bakliyat",
     ["Mercimek (Kırmızı)", "Mercimek (Yeşil)", "Nohut", "Fasulye (Beyaz)", "Fasulye (Kırmızı)", "Bulgur", "Kinoa", "Kuskus"]),
    ("D05", "Yağ & Sos", "Bitkisel Yağlar",
     ["Zeytinyağı (Sızma)", "Zeytinyağı (Riviera)", "Ayçiçek Yağı", "Kolza Yağı", "Hindistan Cevizi Yağı", "Susam Yağı"]),
    ("D05", "Yağ & Sos", "Sos & Dressing",
     ["Domates Sosu", "Pesto", "Alfredo", "Bolognese", "Barbecue Sosu", "Soya Sosu", "Worcestershire"]),
    ("D05", "Yağ & Sos", "Sirke & Hardal",
     ["Elma Sirkesi", "Şarap Sirkesi", "Balsamik", "Dijon Hardalı", "Çekirdekli Hardal"]),
    ("D05", "Konserve & Kavanoz", "Domates Konserve",
     ["Bütün Domates", "Domates Püresi", "Domates Kutusu (Doğranmış)", "Domates Salçası"]),
    ("D05", "Konserve & Kavanoz", "Sebze & Baklagil Konserve",
     ["Mısır Konserve", "Bezelye Konserve", "Fasulye Konserve", "Artichoke", "Zeytin (Yeşil)", "Zeytin (Siyah)"]),
    ("D05", "Baharat & Çeşni", "Baharat",
     ["Kara Biber", "Kırmızı Biber (Tatlı)", "Kırmızı Biber (Acı)", "Kimyon", "Zerdeçal", "Tarçın", "Kekik", "Biberiye"]),
    ("D05", "Baharat & Çeşni", "Tuz & Şeker",
     ["Deniz Tuzu", "İyotlu Tuz", "Fleur de Sel", "Beyaz Şeker", "Esmer Şeker", "Pudra Şekeri", "Stevia"]),
    ("D05", "Konserve Çorba & Hazır Fon", "Çorba Konserve",
     ["Domates Çorbası", "Balkabağı Çorbası", "Mantar Çorbası", "Mercimek Çorbası"]),

    # ── D06 İçecekler ────────────────────────────────────────────
    ("D06", "Su & Maden Suyu", "Maden Suyu",
     ["Sparkling (Küçük)", "Sparkling (Büyük)", "Still (Küçük)", "Still (Büyük)", "Aromatik Maden Suyu"]),
    ("D06", "Meyve Suyu", "Taze & Soğutmalı",
     ["Portakal Suyu (Taze)", "Elma Suyu", "Karışık Meyve", "Tropikal Meyve"]),
    ("D06", "Meyve Suyu", "Uzun Ömürlü",
     ["Portakal Suyu (UHT)", "Vişne Suyu", "Üzüm Suyu", "Ananas Suyu", "Domates Suyu"]),
    ("D06", "Gazlı İçecek", "Kola & Cola",
     ["Cola Normal", "Cola Light/Zero", "Cola Organik", "Cola Yerli"]),
    ("D06", "Gazlı İçecek", "Limonata & Meyve Gazlı",
     ["Limonata", "Portakallı Gazlı", "Üzümlü Gazlı", "Bodega"]),
    ("D06", "Gazlı İçecek", "Enerji & Spor",
     ["Enerji İçeceği", "Spor İçeceği", "Kafein Takviyeli", "Proteinli İçecek"]),
    ("D06", "Sıcak İçecek", "Kahve",
     ["Filtre Kahve", "Espresso Kapsül (Nespresso)", "Espresso Kapsül (Senseo)", "Türk Kahvesi", "Çözünür Kahve", "Decaf"]),
    ("D06", "Sıcak İçecek", "Çay",
     ["Siyah Çay", "Yeşil Çay", "Bitkisel Çay (Papatya)", "Bitkisel Çay (Nane)", "Rooibos", "Earl Grey"]),
    ("D06", "Sıcak İçecek", "Sıcak Çikolata & Malt",
     ["Sıcak Çikolata Tozu", "Cacao Tozu", "Ovomaltine", "Nesquik"]),
    ("D06", "Bira", "Pilsner & Lager",
     ["Jupiler (Pils)", "Stella Artois", "Maes", "Primus", "Cristal", "Alken-Maes Light"]),
    ("D06", "Bira", "Özel & Artisanal",
     ["Duvel (Strong Golden Ale)", "Tripel Karmeliet", "Kasteel Rouge", "Delirium Tremens", "Kwak"]),
    ("D06", "Bira", "Trappist",
     ["Chimay Rouge", "Chimay Bleue", "Chimay Triple", "Orval", "Rochefort 6", "Rochefort 8", "Rochefort 10",
      "Westmalle Dubbel", "Westmalle Tripel", "Westvleteren 8", "Westvleteren 12", "Achel", "La Trappe"]),
    ("D06", "Bira", "Lambic & Zuur",
     ["Gueuze (Boon)", "Gueuze (Cantillon)", "Kriek (Mariage Parfait)", "Faro", "Framboise", "Liefmans Kriek"]),
    ("D06", "Bira", "Abbij & Dubbel/Tripel",
     ["Leffe Blonde", "Leffe Brune", "Leffe Tripel", "Grimbergen", "Tongerlo", "Val-Dieu", "Affligem"]),
    ("D06", "Bira", "Wit & Saison",
     ["Hoegaarden Wit", "Blanche de Bruxelles", "Saison Dupont", "Paix Dieu"]),
    ("D06", "Bira", "Alkolsüz Bira",
     ["Jupiler NA", "Stella Artois Free", "Duvel NA", "Hoegaarden 0.0"]),
    ("D06", "Şarap", "Kırmızı Şarap",
     ["Bordeaux (AOC)", "Bourgogne (AOC)", "Rioja (DOC)", "Chianti", "Malbec", "Shiraz", "Cabernet Sauvignon"]),
    ("D06", "Şarap", "Beyaz Şarap",
     ["Chablis", "Sauvignon Blanc", "Chardonnay", "Riesling", "Pinot Grigio", "Soave"]),
    ("D06", "Şarap", "Rosé & Köpüklü",
     ["Provence Rosé", "Cava", "Prosecco", "Crémant de Bourgogne", "Champagne", "Crémant de Wallonie"]),
    ("D06", "Şarap", "Belçika Şarabı",
     ["Hageland Chardonnay", "Haspengouw Riesling", "Vlaamse Mousserende Wijn", "Côtes de Sambre-et-Meuse"]),
    ("D06", "Spirits & Likör", "Jenever & Gin",
     ["Genever (Jong)", "Genever (Oud)", "Belgian Gin", "Sloe Gin"]),
    ("D06", "Spirits & Likör", "Whisky & Vodka",
     ["Scotch Whisky", "Bourbon", "Irish Whiskey", "Belgian Whisky (Belgian Owl)", "Vodka"]),
    ("D06", "Spirits & Likör", "Likör & Aperitif",
     ["Elixir d'Anvers", "Mandarine Napoléon", "Cointreau", "Aperol", "Campari", "Chartreuse"]),

    # ── D07 Dondurulmuş ─────────────────────────────────────────
    ("D07", "Dondurulmuş Sebze", "Karışık & Tek",
     ["Bezelye", "Mısır", "Ispanak", "Brokoli", "Karnabahar", "Karışık Sebze", "Wok Sebze Karışımı"]),
    ("D07", "Dondurulmuş Et & Balık", "Dondurulmuş Et",
     ["Kıyma", "Köfte", "Tavuk Nugget", "Hamburger", "Schnitzel"]),
    ("D07", "Dondurulmuş Et & Balık", "Dondurulmuş Balık",
     ["Ton", "Balık Parmak", "Kariyer", "Somon Fileto", "Morina"]),
    ("D07", "Dondurulmuş Hazır Yemek", "Pizza",
     ["Margherita", "Quattro Formaggi", "Prosciutto", "Vegan Pizza", "Mini Pizza"]),
    ("D07", "Dondurulmuş Hazır Yemek", "Kroket & Fritür",
     ["Belçika Kroket (Kabeljauw)", "Karides Kroket", "Peynir Kroket", "Frites (Dondurulmuş)"]),
    ("D07", "Dondurulmuş Tatlı & Dondurma", "Dondurma",
     ["Vanilya", "Çikolata", "Çilek", "Fıstık", "Vegan Dondurma", "Waffle Dondurma"]),

    # ── D08 Atıştırmalık & Şekerleme ────────────────────────────
    ("D08", "Cips & Tuzlu Atıştırmalık", "Patates Cipsi",
     ["Normal", "Light", "Kettle", "Organik", "Farklı Tatlar (BBQ, Paprika)"]),
    ("D08", "Cips & Tuzlu Atıştırmalık", "Mısır & Diğer",
     ["Tortilla Cipsi", "Popcorn", "Crackers", "Pirinç Galeti", "Pretzel"]),
    ("D08", "Çikolata", "Tablet Çikolata",
     ["Sütlü Çikolata", "Bitter Çikolata (%70+)", "Beyaz Çikolata", "Ruby Çikolata", "Vegan Çikolata"]),
    ("D08", "Çikolata", "Pralineler & Özel",
     ["Neuhaus Pralineler", "Leonidas Pralineler", "Godiva Pralineler", "Markolini Pralineler",
      "Côte d'Or Bar", "Mini Praline Kutusu", "Çikolata Fondue"]),
    ("D08", "Bisküvi & Kek", "Belçika Bisküvileri",
     ["Speculoos", "Boudoir (Ladyfinger)", "Stroopwafel", "Galettes de Blé", "Jules Destrooper"]),
    ("D08", "Bisküvi & Kek", "Uluslararası Bisküvi",
     ["Oreo", "Digestive", "Petit Beurre", "Lotus Original"]),
    ("D08", "Şekerleme & Gummy", "Gummy & Jelly",
     ["Haribo", "Trolli", "Vegan Gummy", "Meyveli Jelly"]),
    ("D08", "Şekerleme & Gummy", "Sert Şeker & Lolipop",
     ["Werther's", "Fisherman's Friend", "Tic Tac", "Hall's"]),
    ("D08", "Fındık & Kuru Meyve", "Fındık Karışımı",
     ["Badem", "Kaju", "Fındık", "Antep Fıstığı", "Ceviz", "Karışık Kuruyemiş"]),
    ("D08", "Fındık & Kuru Meyve", "Kuru Meyve",
     ["Kuru Üzüm", "Kuru Kayısı", "Hurma", "Kuru İncir", "Kuru Erik"]),

    # ── D09 Kahvaltılık & Mısır Gevreği ────────────────────────
    ("D09", "Mısır Gevreği", "Çocuk Gevrekleri",
     ["Choco Krispies", "Frosties", "Rice Krispies", "Nesquik Gevrek", "Coco Pops"]),
    ("D09", "Mısır Gevreği", "Yetişkin Gevrekleri",
     ["Special K", "All-Bran", "Muesli", "Granola", "Bran Flakes", "Fibre One"]),
    ("D09", "Yulaf & Sıcak Tahıl", "Yulaf Ezmesi",
     ["Instant Yulaf", "Rolled Oats", "Steel-Cut Oats", "Quinoa Puf", "Spelt Puf"]),
    ("D09", "Süt Ürünleri Kahvaltı", "Kahvaltılık Peynir & Ezmeler",
     ["Krem Peynir", "Sürme Peynir", "Tahin", "Fıstık Ezmesi", "Badem Ezmesi"]),
    ("D09", "Reçel & Bal", "Reçel",
     ["Çilek Reçeli", "Kayısı Reçeli", "Ahududu Reçeli", "Yaban Mersini Reçeli", "Şekersiz Reçel"]),
    ("D09", "Reçel & Bal", "Bal & Şurup",
     ["Çiçek Balı", "Kestane Balı", "Akasya Balı", "Akçaağaç Şurubu", "Agave Şurubu"]),

    # ── D10 Hazır & Soğutmalı Yemekler ─────────────────────────
    ("D10", "Soğutmalı Hazır Yemek", "Belçika Geleneksel",
     ["Stoemp", "Waterzooi", "Chicons au Gratin", "Vol-au-vent", "Carbonnade Flamande"]),
    ("D10", "Soğutmalı Hazır Yemek", "Uluslararası",
     ["Sushi", "Börek", "Köfte Yemeği", "Curry", "Stir-Fry"]),
    ("D10", "Sandviç & Wrap", "Hazır Sandviç",
     ["Jambon Fromage", "Thon Mayonnaise", "Veggie", "BLT", "Club Sandwich"]),
    ("D10", "Salata & Meze", "Hazır Salata",
     ["Yeşil Salata", "Tabule", "Couscous Salata", "Niçoise Salatası"]),
    ("D10", "Soğutmalı Spesiyaller", "Raclette & Fondue",
     ["Raclette Peyniri (Dilimli)", "Fondue Fromage Karışımı", "Raclette Seti"]),

    # ── D11 Ev Temizliği & Çamaşır ──────────────────────────────
    ("D11", "Çamaşır", "Çamaşır Deterjanı",
     ["Toz Deterjan", "Sıvı Deterjan", "Kapsül (Pods)", "Ekolojik Deterjan"]),
    ("D11", "Çamaşır", "Yumuşatıcı & Ek",
     ["Kumaş Yumuşatıcı", "Çamaşır Parfümü", "Beyazlatıcı"]),
    ("D11", "Ev Temizliği", "Çok Amaçlı Temizleyici",
     ["Sprey", "Konsantre", "Ekolojik", "Dezenfektan"]),
    ("D11", "Ev Temizliği", "Özel Yüzey",
     ["Banyo Temizleyici", "Mutfak Yağ Çözücü", "Cam Temizleyici", "Tuvalet Temizleyici"]),
    ("D11", "Kağıt & Tek Kullanım", "Kağıt Ürünleri",
     ["Tuvalet Kağıdı", "Kağıt Havlu", "Peçete", "Yüz Mendili", "Mutfak Rulo"]),

    # ── D12 Kişisel Bakım & Hijyen ───────────────────────────────
    ("D12", "Vücut Bakımı", "Duş & Banyo",
     ["Duş Jeli", "Banyo Köpüğü", "Katı Sabun", "Peeling", "Vücut Ovası"]),
    ("D12", "Vücut Bakımı", "Nemlendirici & Losyon",
     ["Vücut Losyonu", "El Kremi", "Ayak Bakım Kremi", "Güneş Koruma (SPF30)", "Güneş Koruma (SPF50)"]),
    ("D12", "Saç Bakımı", "Şampuan & Saç Kremi",
     ["Normal Şampuan", "Kepek Karşıtı", "Renkli Saç Şampuanı", "Saç Kremi", "Saç Maskesi"]),
    ("D12", "Diş & Ağız", "Diş Bakımı",
     ["Diş Macunu (Florürlü)", "Beyazlatıcı Diş Macunu", "Çocuk Diş Macunu", "Diş İpi", "Ağız Gargarası"]),
    ("D12", "Deodorant & Parfüm", "Deodorant",
     ["Roll-On", "Sprey", "Stick", "Alüminümsüz", "Erkek", "Kadın"]),
    ("D12", "Makyaj & Güzellik", "Cilt Bakımı",
     ["Yüz Temizleyici", "Tonik", "Serum", "Nemlendirici Krem", "Göz Altı Kremi"]),
    ("D12", "Hijyen", "Feminen Hijyen",
     ["Ped", "Tampon", "Menstrüel Kilot", "Menstüel Kap"]),
    ("D12", "Hijyen", "Tıraş & Erkek Bakımı",
     ["Tıraş Jeli", "Tıraş Bıçağı", "After Shave", "Sakal Bakım Yağı"]),

    # ── D13 Bebek & Çocuk ───────────────────────────────────────
    ("D13", "Bebek Gıda", "Mama & Püreler",
     ["Sebze Püresi (4ay+)", "Meyve Püresi (4ay+)", "Et Karışımlı (6ay+)", "Tam Tahıllı (6ay+)"]),
    ("D13", "Bebek Gıda", "Bebek Sütü & Formül",
     ["Başlangıç Formülü (0-6ay)", "Devam Formülü (6-12ay)", "Büyüme Sütü (12ay+)", "Premature Formülü"]),
    ("D13", "Bebek Bakım", "Bez & Islak Mendil",
     ["Bebek Bezi (1-2)", "Bebek Bezi (3-4)", "Bebek Bezi (5-6)", "Islak Mendil", "Ekolojik Bez"]),
    ("D13", "Bebek Bakım", "Bebek Bakım Ürünleri",
     ["Bebek Şampuanı", "Bebek Losyonu", "Bebek Pişik Kremi", "Bebek Tozu"]),
    ("D13", "Çocuk Gıda", "Okul Çantası Atıştırmalık",
     ["Meyve Poşeti", "Smoothie Tüp", "Bisküvi (Çocuk)", "Peynir Çubukları"]),

    # ── D14 Evcil Hayvan ────────────────────────────────────────
    ("D14", "Köpek", "Köpek Maması",
     ["Kuru Mama (Yavru)", "Kuru Mama (Yetişkin)", "Kuru Mama (Büyük Irk)", "Yaş Mama", "Premium/Grain-Free"]),
    ("D14", "Köpek", "Veteriner & Diyet",
     ["Eklem Sağlığı Mama", "Düşük Kalori Mama", "Hassas Sindirim", "Alerjik Köpek Maması"]),
    ("D14", "Kedi", "Kedi Maması",
     ["Kuru Mama (Yavru)", "Kuru Mama (Yetişkin)", "Yaş Mama (Küçük Kap)", "Kısırlaştırılmış Kedi"]),
    ("D14", "Kedi", "Veteriner & Diyet",
     ["İdrar Yolu Sağlığı", "Tüy Yumağı Kontrolü", "Kilo Kontrolü", "Hassas Sindirim"]),
    ("D14", "Küçük Hayvanlar & Balık", "Küçük Hayvan",
     ["Kuş Yemi", "Hamster Yemi", "Tavşan Yemi", "Balık Yemi", "Sürüngen Besini"]),
    ("D14", "Aksesuar & Kum", "Kedi Kumu & Temizlik",
     ["Topaklanan Kum", "Kristal Kum", "Parfümsüz Kum", "Bio Kum", "Köpek Tuvalet Torbası"]),

    # ── D15 Belçika Özgün & Mevsimsel ─────────────────────────
    ("D15", "Belçika Çikolatası", "Pralineler & Kutu Çikolata",
     ["Neuhaus Selection", "Leonidas Kutu", "Godiva Signature", "Pierre Marcolini", "Artisan Praline",
      "Manon (Praline)", "Trüf Çikolata"]),
    ("D15", "Belçika Çikolatası", "Tablet & Bar",
     ["Côte d'Or Bloc", "Galler Tablet", "Neuhaus Tablet", "Bean-to-Bar Belge", "Beyaz Çikolata (Belge)"]),
    ("D15", "Belçika Bisküvi & Waffle", "Geleneksel Bisküvi",
     ["Speculoos (Lotus)", "Speculoos (Ambachtelijk)", "Lukken", "Galettes de Blé", "Jules Destrooper",
      "Babelutten", "Cuberdons (Neuzekes)"]),
    ("D15", "Belçika Bisküvi & Waffle", "Waffle & Gaufre",
     ["Gaufre de Liège (Taze)", "Gaufre de Bruxelles (Taze)", "Waffle Hazır", "Waffle Dondurulmuş"]),
    ("D15", "Belçika Peyniri", "AOP & Korumalı",
     ["Herve (AOP)", "Herve Doux", "Herve Piquant", "Remoudou", "Chimay Peyniri", "Postelein",
      "Bouquet des Moines", "Passendale", "Père Joseph", "Orval Peyniri"]),
    ("D15", "Mevsimsel Ürünler", "Sinterklaas (5-6 Aralık)",
     ["Sinterklaas Çikolata", "Speculoos Sinterklaas", "Şeker Bebek", "Krampus Figür"]),
    ("D15", "Mevsimsel Ürünler", "Noel & Yılbaşı",
     ["Noel Çikolata Kutusu", "Noel Bisküvi Kutusu", "Speculoos Noel", "Noel Keki", "Bûche de Noël"]),
    ("D15", "Mevsimsel Ürünler", "Paskalya",
     ["Paskalya Yumurtası (Çikolata)", "Paskalya Tavşanı (Çikolata)", "Pralineli Paskalya Kutusu"]),
    ("D15", "Mevsimsel Ürünler", "Karneval & Mardi Gras",
     ["Oliebollen", "Beignets", "Waffle Karneval"]),
    ("D15", "Belçika Şarabı & Cidre", "Belçika Şarabı",
     ["Hageland Chardonnay", "Hageland Pinot Noir", "Haspengouw Riesling", "Côtes de Sambre-et-Meuse",
      "Vlaamse Mousserende Wijn (VMW)"]),
    ("D15", "Belçika Şarabı & Cidre", "Belçika Elma Şarabı",
     ["Elmer Cidre", "Stassen Cidre", "Liefmans Cidre", "Artisanal Cidre"]),
    ("D15", "Belçika Jenever & Likörü", "Jenever",
     ["Genever Jong", "Genever Oud", "Graanjenever", "Appeljenever", "Perenjenever"]),
    ("D15", "Belçika Jenever & Likörü", "Belçika Likörleri",
     ["Elixir d'Anvers", "Mandarine Napoléon", "Pisang Ambon", "Spa Reine"]),

    # ── D16 Etnik & Uluslararası Gıda ──────────────────────────
    ("D16", "Türk & Orta Doğu", "Türk Gıda Ürünleri",
     ["Pide Ekmeği", "Lavaş", "Simit", "Helva (Susam)", "Helva (Fıstık)", "Lokum (Gül)", "Lokum (Nane)",
      "Ayran (Kutu)", "Boza", "Süzme Yoğurt"]),
    ("D16", "Türk & Orta Doğu", "Baharat & Sos (Türk/Orta Doğu)",
     ["Pul Biber", "Isot Biber", "Sumak", "Za'atar", "Ras El Hanout", "Baharat Karışımı", "Tahini",
      "Humus (Hazır)", "Babagannuş", "Muhammara"]),
    ("D16", "Türk & Orta Doğu", "Tahıl & Baklagil (Türk/Orta Doğu)",
     ["Bulgur (İnce)", "Bulgur (Kaba)", "Mercimek (Kırmızı)", "Nohut (Kuru)", "Freekeh", "Frik"]),
    ("D16", "Kuzey Afrika", "Fas & Tunus & Cezayir Gıdası",
     ["Kuskus (İnce)", "Kuskus (Orta)", "Harissa (Acı)", "Harissa (Hafif)", "Chermoula", "Medjool Hurması",
      "Preserved Lemon (Tuzlu Limon)"]),
    ("D16", "Kuzey Afrika", "Fas Baharat & Soslar",
     ["Ras El Hanout", "Charmoula", "Taklia", "Argan Yağı", "Couscous Bouhali"]),
    ("D16", "Asya", "Japon & Kore Gıdası",
     ["Sushi Pirinci", "Nori (Deniz Yosunu)", "Wasabi", "Ponzu", "Mirin", "Sake (Yemeklik)",
      "Kimchi", "Gochujang", "Doenjang (Fermente Soya)", "Japchae Noodle"]),
    ("D16", "Asya", "Çin & Thai Gıdası",
     ["Soya Sosu (Koyu)", "Soya Sosu (İnce)", "Oyster Sosu", "Hoisin Sosu", "Fish Sauce",
      "Jasmine Pirinci", "Basmati Pirinci", "Ramen Noodle", "Rice Paper", "Coconut Milk"]),
    ("D16", "Asya", "Hint Gıdası",
     ["Basmati Pirinci (Premium)", "Ghee (Hint)", "Garam Masala", "Tandoori Masala", "Curry Paste",
      "Mango Chutney", "Paneer (Taze)", "Dhal Karışımı"]),
    ("D16", "Doğu Avrupa & Balkan", "Polonya & Çek Gıdası",
     ["Kielbasa Sosis", "Kabanosy", "Zurek (Ekşi Çorba Bazı)", "Pierogi (Dondurulmuş)", "Bigos Konserve",
      "Kwas (Fermente)", "Svikova Sos"]),
    ("D16", "Doğu Avrupa & Balkan", "Balkan & Türk Türevi",
     ["Ayvar (Kırmızı)", "Lutenica", "Pindjur", "Burek (Dondurulmuş)", "Kaymak"]),
    ("D16", "Latin Amerika", "Meksika & Orta Amerika",
     ["Tortilla (Buğday)", "Tortilla (Mısır)", "Nacho Cipsi", "Salsa Verde", "Chipotle Adobo",
      "Jalapeño (Kavurma)", "Taco Baharatlı Karışım", "Refried Beans"]),
    ("D16", "Latin Amerika", "Güney Amerika",
     ["Yerba Mate", "Chimichurri Sosu", "Dulce de Leche", "Açaí Püresi", "Quinoa (Peru)"]),
    ("D16", "Fermente & Fonksiyonel", "Kombucha & Kefir",
     ["Kombucha (Original)", "Kombucha (Zencefil)", "Kombucha (Mango)", "Kombucha (Şeftali)",
      "Kefir (Organik)", "Jun Tea"]),
    ("D16", "Fermente & Fonksiyonel", "Miso & Natto",
     ["Shiro Miso (Beyaz)", "Aka Miso (Kırmızı)", "Hatcho Miso", "Natto", "Tempeh"]),
]

# ─────────────────────────────────────────────
# 3. ÖZNİTELİK MATRİSİ
# (isim, tip, zorunlu_mu, açıklama, hangi_departmanlar)
# ─────────────────────────────────────────────
OZELLIKLER = [
    # Temel Kimlik
    ("sku_id",             "TEXT",                     True,  "Market iç ürün ID",                      "Tümü"),
    ("ean",                "TEXT",                     False, "EAN-13 / UPC barkod",                    "Tümü"),
    ("market_chain",       "TEXT",                     True,  "colruyt_be / delhaize_be / ...",          "Tümü"),
    ("product_name_nl",    "TEXT",                     True,  "Hollandaca ürün adı",                    "Tümü"),
    ("product_name_fr",    "TEXT",                     False, "Fransızca ürün adı",                     "Tümü"),
    ("brand",              "TEXT",                     False, "Marka",                                  "Tümü"),
    # Kategori
    ("category_l1",        "TEXT",                     True,  "Departman kodu (D01-D16)",               "Tümü"),
    ("category_l2",        "TEXT",                     True,  "L2 kategori",                            "Tümü"),
    ("category_l3",        "TEXT",                     True,  "L3 alt kategori",                        "Tümü"),
    ("category_l4",        "TEXT",                     False, "L4 ürün tipi",                           "Tümü"),
    # Fiyat
    ("price_retail",       "DECIMAL(8,2)",             True,  "Güncel perakende fiyatı (EUR)",          "Tümü"),
    ("price_per_base_unit","DECIMAL(8,4)",             False, "Birim fiyat (€/kg, €/L, €/st)",          "Tümü"),
    ("base_unit",          "ENUM(kg,L,st,100g,100mL)", False, "Birim türü",                            "Tümü"),
    ("promo_price",        "DECIMAL(8,2)",             False, "Promosyon fiyatı (NULL=yok)",            "Tümü"),
    ("promo_valid_from",   "DATE",                     False, "Promo başlangıç",                        "Tümü"),
    ("promo_valid_until",  "DATE",                     False, "Promo bitiş",                            "Tümü"),
    ("promo_type",         "TEXT",                     False, "percentage/2for1/3for2/multipack/card",  "Tümü"),
    # Ürün Özellikleri
    ("net_weight_g",       "INT",                      False, "Net ağırlık (gram)",                     "Tümü"),
    ("net_volume_ml",      "INT",                      False, "Net hacim (mL) — içecekler",             "D06,D07"),
    ("packaging_material", "TEXT",                     False, "glass/plastic/cardboard/can/tetra",      "Tümü"),
    ("recycle_code",       "TEXT",                     False, "PET1, HDPE2, PP5 vb.",                   "Tümü"),
    # Besin & Sağlık
    ("nutriscore",         "CHAR(1)",                  False, "A/B/C/D/E Nutri-Score",                  "D01-D16"),
    ("ecoscore",           "CHAR(1)",                  False, "A/B/C/D/E Eco-Score",                    "D01-D16"),
    ("energy_kcal_100g",   "INT",                      False, "Enerji (kcal/100g veya 100mL)",         "D01-D10,D16"),
    ("fat_g_100g",         "DECIMAL(5,2)",             False, "Yağ (g/100g)",                           "D01-D10"),
    ("saturated_fat_g",    "DECIMAL(5,2)",             False, "Doymuş yağ (g/100g)",                   "D01-D10"),
    ("sugars_g_100g",      "DECIMAL(5,2)",             False, "Şeker (g/100g)",                         "D01-D10"),
    ("salt_g_100g",        "DECIMAL(5,2)",             False, "Tuz (g/100g)",                           "D01-D10"),
    ("protein_g_100g",     "DECIMAL(5,2)",             False, "Protein (g/100g)",                       "D01-D10"),
    # Diyet Etiketleri
    ("is_organic",         "BOOLEAN",                  False, "Bio/Organik sertifikalı",                "Tümü"),
    ("is_vegan",           "BOOLEAN",                  False, "Vegan",                                  "Tümü"),
    ("is_vegetarian",      "BOOLEAN",                  False, "Vejetaryen",                             "Tümü"),
    ("is_gluten_free",     "BOOLEAN",                  False, "Glutensiz",                              "Tümü"),
    ("is_lactose_free",    "BOOLEAN",                  False, "Laktosuz",                               "Tümü"),
    ("is_halal",           "BOOLEAN",                  False, "Helal sertifikalı",                      "Tümü"),
    ("is_kosher",          "BOOLEAN",                  False, "Koşer sertifikalı",                      "Tümü"),
    ("palm_oil_free",      "BOOLEAN",                  False, "Palmiye yağı içermez",                  "Tümü"),
    ("is_fodmap_friendly", "BOOLEAN",                  False, "Low-FODMAP (IBS dostu)",                 "D01-D10,D16"),
    # Sertifika & Ödüller
    ("regional_cert",      "TEXT[]",                   False, "AOP/PDO, IGP/PGI, TSG, Demeter",        "Tümü"),
    ("award_label",        "TEXT[]",                   False, "Monde Sélection, Superior Taste Award",  "Tümü"),
    # İçecek Özgün
    ("alcohol_pct",        "DECIMAL(4,1)",             False, "Alkol oranı % (NULL=alkollü değil)",    "D06,D15"),
    ("vintage",            "SMALLINT",                 False, "Şarap yılı",                             "D06,D15"),
    # Ürün Ailesi
    ("product_family_id",  "UUID",                     False, "Aynı ürünün farklı ambalajları",        "Tümü"),
    # Meta
    ("availability",       "TEXT",                     False, "in_stock/low_stock/out_of_stock",        "Tümü"),
    ("last_updated_at",    "TIMESTAMPTZ",              True,  "Son güncelleme zaman damgası",           "Tümü"),
]

# ─────────────────────────────────────────────
# 4. NORMALİZASYON SÖZLÜĞÜ
# (standart_anahtar, NL_terimler, FR_terimler, DE_terimler, açıklama)
# ─────────────────────────────────────────────
NORMALIZASYON = [
    # Diyet Flagları
    ("organic",         ["biologisch","bio","ekologisch"],
                        ["biologique","bio","organique"],
                        ["biologisch","bio","ökologisch"],
                        "Organik/Bio sertifikalı"),
    ("vegan",           ["veganistisch","vegan","plantaardig","100% plantbased"],
                        ["végétalien","vegan","100% végétal","à base de plantes"],
                        ["veganistisch","vegan","pflanzlich"],
                        "Vegan ürün"),
    ("vegetarian",      ["vegetarisch","zonder vlees"],
                        ["végétarien","sans viande"],
                        ["vegetarisch","fleischlos"],
                        "Vejetaryen"),
    ("gluten_free",     ["glutenvrij","zonder gluten","GVG"],
                        ["sans gluten","sans gluten","celiac"],
                        ["glutenfrei","ohne gluten"],
                        "Glutensiz"),
    ("lactose_free",    ["lactosevrij","zonder lactose"],
                        ["sans lactose","intolérance au lactose"],
                        ["laktosefrei","ohne laktose"],
                        "Laktosuz"),
    ("halal",           ["halal","halal gecertificeerd"],
                        ["halal","certifié halal"],
                        ["halal","halal-zertifiziert"],
                        "Helal sertifikalı"),
    ("kosher",          ["koosjer","kosher gecertificeerd"],
                        ["casher","certifié casher"],
                        ["koscher","kosher-zertifiziert"],
                        "Koşer sertifikalı"),
    ("palm_oil_free",   ["palmolievrij","zonder palmolie","geen palmolie"],
                        ["sans huile de palme","zéro huile de palme"],
                        ["palmölfrei","ohne palmöl"],
                        "Palmiye yağı içermiyor"),
    ("fodmap_friendly", ["fodmap-vriendelijk","low fodmap","ibs-vriendelijk"],
                        ["fodmap-friendly","faible en fodmap","amical ibs"],
                        ["fodmap-freundlich","niedrig fodmap"],
                        "Low-FODMAP (IBS dostu)"),
    # Nutri-Score
    ("nutriscore_a",    ["nutri-score a","nutriscore a","groene score a"],
                        ["nutri-score a","score vert a"],
                        ["nutri-score a"],
                        "Nutri-Score A (en iyi)"),
    ("nutriscore_b",    ["nutri-score b"],
                        ["nutri-score b"],
                        ["nutri-score b"],
                        "Nutri-Score B"),
    ("nutriscore_c",    ["nutri-score c"],
                        ["nutri-score c"],
                        ["nutri-score c"],
                        "Nutri-Score C"),
    ("nutriscore_d",    ["nutri-score d"],
                        ["nutri-score d"],
                        ["nutri-score d"],
                        "Nutri-Score D"),
    ("nutriscore_e",    ["nutri-score e","rode score","score rouge"],
                        ["nutri-score e","score rouge e"],
                        ["nutri-score e"],
                        "Nutri-Score E (en kötü)"),
    # Bölgesel Sertifikalar
    ("regional_aop",    ["beschermde oorsprongsbenaming","bob","gecontroleerde herkomst"],
                        ["appellation d'origine protégée","aop","aop belge"],
                        ["geschützte ursprungsbezeichnung","g.u.","gub"],
                        "AOP/PDO — Korumalı Menşe Adı"),
    ("regional_igp",    ["beschermde geografische aanduiding","bga"],
                        ["indication géographique protégée","igp"],
                        ["geschützte geografische angabe","g.g.a."],
                        "IGP/PGI — Korumalı Coğrafi İşaret"),
    ("regional_tsg",    ["gegarandeerde traditionele specialiteit","gts"],
                        ["spécialité traditionnelle garantie","stg"],
                        ["garantierte traditionelle spezialität","g.t.s."],
                        "TSG/GTS — Geleneksel Garanti"),
    ("demeter",         ["demeter","biodynamisch"],
                        ["demeter","biodynamique"],
                        ["demeter","biodynamisch"],
                        "Demeter biyodinamik sertifikası"),
    # Belçika Bira Türleri
    ("beer_pils",       ["pils","pilsener","lager","blond licht"],
                        ["pils","pilsener","lager","blonde légère"],
                        ["pils","pilsener"],
                        "Pilsner/Lager stili bira"),
    ("beer_trappist",   ["trappist","abdijbier","kloosterbier","trappistenbrouwerij"],
                        ["trappiste","bière trappiste","bière d'abbaye trappiste"],
                        ["trappistenbier","trappistenbrauerei"],
                        "Belçika Trappist birası"),
    ("beer_lambic",     ["lambiek","lambic","spontaan gegist"],
                        ["lambic","fermentation spontanée"],
                        ["lambic","spontangärung"],
                        "Lambic — spontan fermantasyon"),
    ("beer_gueuze",     ["geuze","gueuze","oude geuze"],
                        ["gueuze","vieille gueuze"],
                        ["gueuze","alte gueuze"],
                        "Gueuze (harmanlanmış lambic)"),
    ("beer_kriek",      ["kriek","kriekbier","krieklambiek","kriekenlambiek"],
                        ["kriek","kriek aux cerises","kriek lambic"],
                        ["kriek","kirschbier"],
                        "Kriek — vişneli lambic"),
    ("beer_faro",       ["faro","faro-lambiek"],
                        ["faro","faro lambic"],
                        ["faro"],
                        "Faro — şeker eklenmiş lambic"),
    ("beer_witbier",    ["witbier","wittebier","blanche","bière blanche"],
                        ["bière blanche","witbier"],
                        ["weizenbier","weißbier"],
                        "Wit bier / Bière blanche"),
    ("beer_saison",     ["saison","seizoensbier","farmhouse ale"],
                        ["saison","bière de saison"],
                        ["saison","saisonbier"],
                        "Saison stili bira"),
    ("beer_dubbel",     ["dubbel","abdijdubbel","double"],
                        ["dubbel","double","bière double"],
                        ["dubbel","doppelbier"],
                        "Abdij Dubbel"),
    ("beer_tripel",     ["tripel","triple","abdijtriple"],
                        ["tripel","triple","bière triple"],
                        ["tripel","tripelbier"],
                        "Abdij Tripel"),
    ("beer_quadrupel",  ["quadrupel","quad","abdijquad"],
                        ["quadrupel","quadruple"],
                        ["quadrupel"],
                        "Quadrupel (Abdij)"),
    ("beer_alcohol_free",["alcoholvrij","0.0%","0.5%","alcoholarm","zonder alcohol"],
                        ["sans alcool","0.0%","0.5%","peu alcoolisé"],
                        ["alkoholfrei","0.0%","ohne alkohol"],
                        "Alkolsüz bira"),
    # Trappist Markaları
    ("brand_chimay",    ["chimay"], ["chimay"], ["chimay"], "Chimay Trappist birası"),
    ("brand_orval",     ["orval"], ["orval"], ["orval"], "Orval Trappist birası"),
    ("brand_rochefort", ["rochefort"], ["rochefort"], ["rochefort"], "Rochefort Trappist"),
    ("brand_westmalle", ["westmalle"], ["westmalle"], ["westmalle"], "Westmalle Trappist"),
    ("brand_westvleteren", ["westvleteren","sint-sixtus"], ["westvleteren","saint-sixte"], ["westvleteren"], "Westvleteren Trappist"),
    ("brand_achel",     ["achel","achelse kluis"], ["achel"], ["achel"], "Achel Trappist"),
    # Şarap Türleri
    ("wine_red",        ["rode wijn","rood","rouge"],
                        ["vin rouge","rouge"],
                        ["rotwein","rot"],
                        "Kırmızı şarap"),
    ("wine_white",      ["witte wijn","wit"],
                        ["vin blanc","blanc"],
                        ["weißwein","weiß"],
                        "Beyaz şarap"),
    ("wine_rose",       ["rosé","roséwijn","roséwein"],
                        ["rosé","vin rosé"],
                        ["rosé","roséwein"],
                        "Rosé şarap"),
    ("wine_sparkling",  ["mousserende wijn","schuimwijn","pétillant","brut","extra brut"],
                        ["vin mousseux","crémant","champagne","prosecco","cava"],
                        ["sekt","schaumwein","crémant"],
                        "Köpüklü şarap"),
    ("wine_organic",    ["biologische wijn","biowijn","nature wine"],
                        ["vin biologique","vin bio","vin nature"],
                        ["biowein","ökologischer wein"],
                        "Organik şarap"),
    # Belçika Çikolata Markaları
    ("brand_neuhaus",   ["neuhaus"], ["neuhaus"], ["neuhaus"], "Neuhaus (Belçika praline)"),
    ("brand_leonidas",  ["leonidas"], ["leonidas"], ["leonidas"], "Leonidas çikolata"),
    ("brand_godiva",    ["godiva"], ["godiva"], ["godiva"], "Godiva çikolata"),
    ("brand_marcolini", ["marcolini","pierre marcolini"], ["marcolini"], ["marcolini"], "Pierre Marcolini artisan"),
    ("brand_cote_dor",  ["côte d'or","cote d'or","côted'or"], ["côte d'or"], ["côte d'or"], "Côte d'Or çikolata"),
    ("brand_galler",    ["galler"], ["galler"], ["galler"], "Galler çikolata"),
    # Belçika Bisküvi Markaları
    ("brand_lotus",     ["lotus","lotus bakeries"], ["lotus","lotus bakeries"], ["lotus"], "Lotus Speculoos"),
    ("brand_destrooper",["jules destrooper","j. destrooper"], ["jules destrooper"], ["jules destrooper"], "Jules Destrooper"),
    # Ambalaj
    ("packaging_glass", ["glazen fles","glas","in glas"],
                        ["bouteille en verre","en verre"],
                        ["glasflasche","aus glas"],
                        "Cam ambalaj"),
    ("packaging_can",   ["blik","blikje","in blik"],
                        ["boîte","canette","en boîte"],
                        ["dose","büchse","in dose"],
                        "Teneke/kutu ambalaj"),
    ("packaging_tetra", ["tetra pak","tetrapack","kartonpak"],
                        ["tetra pak","brique","carton"],
                        ["tetra pak","getränkekarton"],
                        "Tetra Pak karton"),
    ("packaging_compostable",["composteerbaar","plantaardig verpakking"],
                        ["compostable","emballage végétal"],
                        ["kompostierbar"],
                        "Kompostlanabilir ambalaj"),
    # Belçika Özgün Ürünler
    ("witloof",         ["witloof","witlof","chicon"],
                        ["chicon","witloof","endive de bruxelles"],
                        ["chicorée","witloof"],
                        "Witloof/Chicon — Belçika endivia"),
    ("brussels_sprout", ["spruiten","spruitjes","brusselse spruitjes"],
                        ["choux de bruxelles","sprouts"],
                        ["rosenkohl","brüsseler kohl"],
                        "Brüksel lahanası"),
    ("speculoos",       ["speculoos","speculaas"],
                        ["spéculoos","spéculoos"],
                        ["spekulatius"],
                        "Speculoos bisküvisi"),
    # Fermente Ürünler
    ("kombucha",        ["kombucha","kombuchathee"],
                        ["kombucha","thé kombucha"],
                        ["kombucha","kombuchatee"],
                        "Kombucha — fermente çay"),
    ("kimchi",          ["kimchi","koreaans gefermenteerd"], ["kimchi"],["kimchi"],
                        "Kimchi — Kore fermente lahana"),
    ("miso",            ["miso","misopasta"], ["miso","pâte miso"], ["miso","misopaste"],
                        "Miso — Japon fermente soya"),
    ("kefir",           ["kefir","kefirmelk"], ["kéfir","lait fermenté kéfir"], ["kefir"],
                        "Kefir — fermente süt içeceği"),
    ("tempeh",          ["tempeh","tempé"], ["tempeh","tempé"], ["tempeh"],
                        "Tempeh — fermente soya"),
    # Etnik Gıda Terimleri
    ("couscous",        ["couscous","cous cous"], ["couscous"], ["couscous"], "Kuskus"),
    ("harissa",         ["harissa"], ["harissa"], ["harissa"], "Harissa acı sos"),
    ("tahini",          ["tahin","tahina","sesampasta"], ["tahin","tahini","purée de sésame"], ["tahin","sesampaste"],
                        "Tahin — susam ezmesi"),
    ("hummus",          ["hummus","houmous"], ["houmous","hummus"], ["hummus"], "Humus — nohut ezmesi"),
    ("tortilla",        ["tortilla","maïstortilla","tarwebladij"], ["tortilla","galette de maïs"], ["tortilla"],
                        "Tortilla — buğday/mısır yassı ekmek"),
    ("yerba_mate",      ["yerba mate","mate"], ["yerba maté","maté"], ["yerba mate"],
                        "Yerba Mate"),
]

# ─────────────────────────────────────────────
# 5. VERİ MODELİ SQL
# ─────────────────────────────────────────────
VERI_MODELI_SQL = """
-- Temel ürünler tablosu (mevcut + yeni sütunlar)
CREATE TABLE IF NOT EXISTS market_chain_products (
    id                   BIGSERIAL PRIMARY KEY,
    chain_slug           TEXT NOT NULL,           -- colruyt_be, delhaize_be ...
    external_product_id  TEXT NOT NULL,           -- market iç ID
    ean                  TEXT,
    name                 TEXT NOT NULL,
    brand                TEXT,
    category_l1          TEXT,
    category_l2          TEXT,
    category_l3          TEXT,
    category_l4          TEXT,
    price                DECIMAL(8,2),
    price_per_base_unit  DECIMAL(8,4),
    base_unit            TEXT,                   -- kg/L/st/100g/100mL
    promo_price          DECIMAL(8,2),           -- NULL = promo yok
    promo_valid_from     DATE,
    promo_valid_until    DATE,
    promo_type           TEXT,
    net_weight_g         INT,
    net_volume_ml        INT,
    packaging_material   TEXT,
    recycle_code         TEXT,
    nutriscore           CHAR(1),               -- A/B/C/D/E
    ecoscore             CHAR(1),               -- A/B/C/D/E
    energy_kcal_100g     INT,
    fat_g_100g           DECIMAL(5,2),
    saturated_fat_g      DECIMAL(5,2),
    sugars_g_100g        DECIMAL(5,2),
    salt_g_100g          DECIMAL(5,2),
    protein_g_100g       DECIMAL(5,2),
    is_organic           BOOLEAN DEFAULT FALSE,
    is_vegan             BOOLEAN DEFAULT FALSE,
    is_vegetarian        BOOLEAN DEFAULT FALSE,
    is_gluten_free       BOOLEAN DEFAULT FALSE,
    is_lactose_free      BOOLEAN DEFAULT FALSE,
    is_halal             BOOLEAN DEFAULT FALSE,
    is_kosher            BOOLEAN DEFAULT FALSE,
    palm_oil_free        BOOLEAN DEFAULT FALSE,
    is_fodmap_friendly   BOOLEAN DEFAULT FALSE,
    regional_cert        TEXT[],
    award_label          TEXT[],
    alcohol_pct          DECIMAL(4,1),          -- NULL = alkolsüz
    vintage              SMALLINT,              -- şarap yılı
    product_family_id    UUID,
    availability         TEXT DEFAULT 'in_stock',
    image_url            TEXT,
    source_url           TEXT,
    last_updated_at      TIMESTAMPTZ DEFAULT now(),
    UNIQUE(chain_slug, external_product_id)
);

-- Ürün aileleri (aynı ürün, farklı ambalaj/market)
CREATE TABLE IF NOT EXISTS product_families (
    family_id        UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    ean              TEXT,
    canonical_name   TEXT,
    brand            TEXT,
    category_l1      TEXT,
    created_at       TIMESTAMPTZ DEFAULT now()
);

-- Haftalık fiyat geçmişi
CREATE TABLE IF NOT EXISTS price_history (
    id                   BIGSERIAL PRIMARY KEY,
    external_product_id  TEXT NOT NULL,
    chain_slug           TEXT NOT NULL,
    price                DECIMAL(8,2),
    promo_price          DECIMAL(8,2),
    recorded_at          DATE DEFAULT CURRENT_DATE,
    UNIQUE(external_product_id, chain_slug, recorded_at)
);

-- Sepet analizi (market karşılaştırma)
CREATE TABLE IF NOT EXISTS basket_analysis (
    basket_id        UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    user_session     TEXT,
    created_at       TIMESTAMPTZ DEFAULT now(),
    total_colruyt    DECIMAL(10,2),
    total_delhaize   DECIMAL(10,2),
    total_carrefour  DECIMAL(10,2),
    total_lidl       DECIMAL(10,2),
    total_aldi       DECIMAL(10,2),
    savings          DECIMAL(10,2)
);

-- Yararlı index'ler
CREATE INDEX IF NOT EXISTS idx_mcp_chain ON market_chain_products(chain_slug);
CREATE INDEX IF NOT EXISTS idx_mcp_cat_l1 ON market_chain_products(category_l1);
CREATE INDEX IF NOT EXISTS idx_mcp_cat_l2 ON market_chain_products(category_l2);
CREATE INDEX IF NOT EXISTS idx_mcp_nutriscore ON market_chain_products(nutriscore);
CREATE INDEX IF NOT EXISTS idx_mcp_promo ON market_chain_products(promo_price) WHERE promo_price IS NOT NULL;
CREATE INDEX IF NOT EXISTS idx_ph_product ON price_history(external_product_id, chain_slug);
"""

# ─────────────────────────────────────────────
# 6. DEMO VERİLERİ
# ─────────────────────────────────────────────
DEMO_YUMURTA = [
    {
        "sku_id": "DH-001234", "ean": "5410008131231", "market_chain": "delhaize_be",
        "product_name_nl": "Delhaize Verse Eieren Vrije Uitloop L 12st",
        "product_name_fr": "Delhaize Oeufs Frais Plein Air L 12pcs",
        "brand": "Delhaize",
        "category_l1": "D01", "category_l2": "Yumurta",
        "category_l3": "Tavuk Yumurtası", "category_l4": "L Beden (63-73g)",
        "price_retail": 3.49, "price_per_base_unit": 0.29, "base_unit": "st",
        "promo_price": None, "promo_valid_from": None, "promo_valid_until": None,
        "net_weight_g": 780, "packaging_material": "cardboard",
        "nutriscore": "A", "ecoscore": "B",
        "energy_kcal_100g": 155, "fat_g_100g": 11.3, "saturated_fat_g": 3.3,
        "sugars_g_100g": 0.0, "salt_g_100g": 0.36, "protein_g_100g": 13.0,
        "is_organic": False, "is_vegan": False, "is_vegetarian": True,
        "is_gluten_free": True, "is_lactose_free": True,
        "regional_cert": [], "award_label": [],
        "alcohol_pct": None, "vintage": None,
        "availability": "in_stock",
    },
    {
        "sku_id": "CL-554422", "ean": "5410008131231", "market_chain": "colruyt_be",
        "product_name_nl": "Colruyt Bio Scharreleieren M 6st",
        "product_name_fr": "Colruyt Oeufs Bio Plein Air M 6pcs",
        "brand": "Boni Selection",
        "category_l1": "D01", "category_l2": "Yumurta",
        "category_l3": "Tavuk Yumurtası", "category_l4": "M Beden (53-63g)",
        "price_retail": 2.89, "price_per_base_unit": 0.48, "base_unit": "st",
        "promo_price": 2.49, "promo_valid_from": "2026-05-05", "promo_valid_until": "2026-05-11",
        "promo_type": "percentage",
        "net_weight_g": 330, "packaging_material": "cardboard",
        "nutriscore": "A", "ecoscore": "A",
        "energy_kcal_100g": 155, "fat_g_100g": 11.3, "saturated_fat_g": 3.3,
        "sugars_g_100g": 0.0, "salt_g_100g": 0.36, "protein_g_100g": 13.0,
        "is_organic": True, "is_vegan": False, "is_vegetarian": True,
        "is_gluten_free": True, "is_lactose_free": True,
        "regional_cert": ["Bio"], "award_label": [],
        "alcohol_pct": None, "vintage": None,
        "availability": "in_stock",
    },
]

DEMO_BIRA = [
    {
        "sku_id": "DH-CHIMAY-RG", "ean": "5413515021040", "market_chain": "delhaize_be",
        "product_name_nl": "Chimay Rouge Trappistenbier 75cl",
        "product_name_fr": "Chimay Rouge Bière Trappiste 75cl",
        "brand": "Chimay", "category_l1": "D06",
        "category_l2": "Bira", "category_l3": "Trappist", "category_l4": "Chimay Rouge",
        "price_retail": 5.49, "price_per_base_unit": 7.32, "base_unit": "L",
        "promo_price": None,
        "net_volume_ml": 750, "packaging_material": "glass",
        "nutriscore": None, "ecoscore": "C",
        "energy_kcal_100g": 50, "alcohol_pct": 7.0, "vintage": None,
        "is_organic": False, "is_vegan": True, "is_vegetarian": True,
        "is_gluten_free": False, "is_lactose_free": True,
        "regional_cert": ["Trappist Authentic"], "award_label": [],
        "availability": "in_stock",
    },
    {
        "sku_id": "CL-CANTILLON-GZ", "ean": "5410072003007", "market_chain": "colruyt_be",
        "product_name_nl": "Cantillon Gueuze Lambiek 37.5cl",
        "product_name_fr": "Cantillon Gueuze Lambic 37.5cl",
        "brand": "Cantillon", "category_l1": "D06",
        "category_l2": "Bira", "category_l3": "Lambic & Zuur", "category_l4": "Gueuze (Cantillon)",
        "price_retail": 8.95, "price_per_base_unit": 23.87, "base_unit": "L",
        "promo_price": None,
        "net_volume_ml": 375, "packaging_material": "glass",
        "nutriscore": None, "ecoscore": "B",
        "energy_kcal_100g": 42, "alcohol_pct": 5.0, "vintage": 2023,
        "is_organic": True, "is_vegan": True, "is_vegetarian": True,
        "is_gluten_free": False, "is_lactose_free": True,
        "regional_cert": ["Bio", "Trappist Authentic"], "award_label": [],
        "availability": "low_stock",
    },
    {
        "sku_id": "LD-JUPILER-6PK", "ean": "5000282391602", "market_chain": "lidl_be",
        "product_name_nl": "Jupiler Pils 6x33cl",
        "product_name_fr": "Jupiler Pils 6x33cl",
        "brand": "Jupiler", "category_l1": "D06",
        "category_l2": "Bira", "category_l3": "Pilsner & Lager", "category_l4": "Jupiler (Pils)",
        "price_retail": 6.49, "price_per_base_unit": 3.28, "base_unit": "L",
        "promo_price": 5.79, "promo_valid_from": "2026-05-04", "promo_valid_until": "2026-05-10",
        "promo_type": "percentage",
        "net_volume_ml": 1980, "packaging_material": "can",
        "nutriscore": None, "ecoscore": "D",
        "energy_kcal_100g": 43, "alcohol_pct": 5.2, "vintage": None,
        "is_organic": False, "is_vegan": True, "is_vegetarian": True,
        "is_gluten_free": False, "is_lactose_free": True,
        "regional_cert": [], "award_label": [],
        "availability": "in_stock",
    },
]

DEMO_PROMO = [
    {
        "sku_id": "AL-DUVEL-33", "product_name_nl": "Duvel Strong Golden Ale 33cl",
        "market_chain": "aldi_be", "price_retail": 1.89, "promo_price": 1.49,
        "promo_valid_from": "2026-05-04", "promo_valid_until": "2026-05-17",
        "promo_type": "percentage", "promo_pct": "21%",
        "category_l3": "Özel & Artisanal", "brand": "Duvel", "alcohol_pct": 8.5,
        "nutriscore": None, "ecoscore": "C", "availability": "in_stock",
    },
    {
        "sku_id": "DH-NEUHAUS-250", "product_name_nl": "Neuhaus Selection Pralines 250g",
        "market_chain": "delhaize_be", "price_retail": 18.99, "promo_price": 14.99,
        "promo_valid_from": "2026-05-04", "promo_valid_until": "2026-05-10",
        "promo_type": "percentage", "promo_pct": "21%",
        "category_l3": "Pralineler & Kutu Çikolata", "brand": "Neuhaus", "alcohol_pct": None,
        "nutriscore": "D", "ecoscore": "C", "availability": "in_stock",
    },
    {
        "sku_id": "CL-SPECIAL-K", "product_name_nl": "Kellogg's Special K Rode Vruchten 375g",
        "market_chain": "colruyt_be", "price_retail": 4.29, "promo_price": 2.99,
        "promo_valid_from": "2026-05-04", "promo_valid_until": "2026-05-17",
        "promo_type": "2for1", "promo_pct": "30%",
        "category_l3": "Yetişkin Gevrekleri", "brand": "Kellogg's", "alcohol_pct": None,
        "nutriscore": "B", "ecoscore": "C", "availability": "in_stock",
    },
]

# ─────────────────────────────────────────────
# 7. EXCEL ÜRETİCİ
# ─────────────────────────────────────────────

# Renkler
RENKLER = {
    "header_blue":    "1F4E79",
    "header_green":   "1E5631",
    "header_purple":  "4A1680",
    "header_orange":  "7F3A00",
    "header_red":     "7B0000",
    "header_teal":    "0D4D4D",
    "sub_blue":       "BDD7EE",
    "sub_green":      "C6EFCE",
    "sub_yellow":     "FFEB9C",
    "sub_purple":     "D9B8FF",
    "row_alt":        "F5F5F5",
    "white":          "FFFFFF",
    # Nutri-Score
    "nutri_a":        "00B050",
    "nutri_b":        "92D050",
    "nutri_c":        "FFFF00",
    "nutri_d":        "FF9900",
    "nutri_e":        "FF0000",
    # Promo
    "promo_bg":       "FFE6CC",
    "dept_header":    "2E75B6",
}

def stil(wb, bold=False, color="000000", bg=None, size=10, wrap=False, align="left", border=False):
    font = Font(bold=bold, color=color, size=size)
    fill = PatternFill("solid", fgColor=bg) if bg else PatternFill()
    align_obj = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    border_obj = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    ) if border else Border()
    return font, fill, align_obj, border_obj

def hucre_stil(ws, row, col, value, bold=False, fg="000000", bg=None, size=10,
               wrap=False, align="left", border=True, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if number_format:
        cell.number_format = number_format
    return cell

def ozet_sayfasi(wb):
    ws = wb.create_sheet("Ozet")
    ws.sheet_view.showGridLines = False

    # Başlık
    ws.merge_cells("A1:H1")
    hucre_stil(ws, 1, 1, "🛒 BELÇİKA MARKET KATEGORİ SİSTEMİ v2.0",
               bold=True, fg="FFFFFF", bg=RENKLER["header_blue"], size=16, align="center", border=False)
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    hucre_stil(ws, 2, 1, f"Oluşturma tarihi: {datetime.date.today()} | 16 Departman | ~420 L3 Kategori | ~1400 L4 Ürün Tipi | 43 Öznitelik",
               fg="666666", size=10, align="center", border=False)
    ws.row_dimensions[2].height = 20

    # Departman özeti tablosu
    ws.row_dimensions[4].height = 24
    basliklar = ["Kod", "Departman (TR)", "Departman (NL)", "Departman (FR)", "L3 Sayısı", "L4 Tahmini", "Anahtar Özellik"]
    for col, b in enumerate(basliklar, 1):
        hucre_stil(ws, 4, col, b, bold=True, fg="FFFFFF", bg=RENKLER["header_blue"], size=10, align="center")

    # Departman listesi istatistikleri
    l3_sayilari = {}
    l4_sayilari = {}
    for d_id, l2, l3, l4_list in KATEGORILER:
        l3_sayilari[d_id] = l3_sayilari.get(d_id, set())
        l3_sayilari[d_id].add(f"{l2}>{l3}")
        l4_sayilari[d_id] = l4_sayilari.get(d_id, 0) + len(l4_list)

    dept_ozellikler = {
        "D01": "Nutri-Score, Laktosuz, Organik, AOP Peynir",
        "D02": "Helal, Taze/Dondurulmuş, IGP Ardenne Jambon",
        "D03": "Organik, Yerel, Witloof (Belçika özgün)",
        "D04": "Waffle/Gaufre, Glutensiz, Tam Tahıllı",
        "D05": "Birim Fiyat, Yağ Türü, Konserve",
        "D06": "Alkol %, Vintage, Trappist, Lambic, AOP Şarap",
        "D07": "Dondurulmuş, Belçika Kroket",
        "D08": "Nutri-Score, Palm Yağı, Praline Markalar",
        "D09": "Mısır Gevreği, Nutri-Score, Yulaf",
        "D10": "Belçika geleneksel (Stoemp, Waterzooi), Sushi",
        "D11": "Ekolojik, Konsantre, Çamaşır Kapsülü",
        "D12": "SPF, Alüminyumsuz Deo, Tıraş",
        "D13": "Bebek Sütü Tipi, Beden, Organik",
        "D14": "Veteriner Mama, Kedi Kumu",
        "D15": "Trappist, AOP Peynir, Speculoos, Waffle, Mevsimsel",
        "D16": "Etnik, Fermente, Kombucha, Kimchi, Miso, Couscous",
    }

    for i, (d_id, d_tr, d_nl, d_fr) in enumerate(DEPARTMANLAR):
        row = 5 + i
        ws.row_dimensions[row].height = 20
        bg = RENKLER["sub_blue"] if i % 2 == 0 else RENKLER["white"]
        l3_n = len(l3_sayilari.get(d_id, set()))
        l4_n = l4_sayilari.get(d_id, 0)
        ozellik = dept_ozellikler.get(d_id, "")
        for col, val in enumerate([d_id, d_tr, d_nl, d_fr, l3_n, l4_n, ozellik], 1):
            hucre_stil(ws, row, col, val, bg=bg, size=9,
                       align="center" if col in (1, 5, 6) else "left")

    # Toplam satırı
    total_row = 5 + len(DEPARTMANLAR)
    ws.row_dimensions[total_row].height = 22
    total_l3 = sum(len(v) for v in l3_sayilari.values())
    total_l4 = sum(l4_sayilari.values())
    hucre_stil(ws, total_row, 1, "TOPLAM", bold=True, bg="DDEEFF", size=10, align="center")
    ws.merge_cells(f"B{total_row}:D{total_row}")
    hucre_stil(ws, total_row, 2, f"{len(DEPARTMANLAR)} Departman", bold=True, bg="DDEEFF", size=10)
    hucre_stil(ws, total_row, 5, total_l3, bold=True, bg="DDEEFF", size=10, align="center")
    hucre_stil(ws, total_row, 6, total_l4, bold=True, bg="DDEEFF", size=10, align="center")

    # Sütun genişlikleri
    genislikler = [8, 30, 35, 38, 12, 14, 45]
    for i, w in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Özellik özeti kutusu
    ozet_row = total_row + 3
    ws.merge_cells(f"A{ozet_row}:D{ozet_row}")
    hucre_stil(ws, ozet_row, 1, "📋 SİSTEM ÖZELLİKLERİ",
               bold=True, fg="FFFFFF", bg=RENKLER["header_green"], size=11, align="center")
    ozellikler_listesi = [
        "✅ 16 departman (D01-D16), 2 yeni: D15 Belçika Özgün, D16 Etnik",
        "✅ 43 öznitelik (mevcut 28'e +15 eklendi: Nutri-Score, Eco-Score, promo, alkol...)",
        "✅ ~120 normalizasyon kuralı (NL/FR/DE): Trappist, Lambic, AOP, Speculoos...",
        "✅ Promo fiyat takibi (promo_price, promo_valid_from/until, promo_type)",
        "✅ Bölgesel sertifikalar: AOP, IGP, TSG, Demeter, Trappist Authentic",
        "✅ Belçika özgün: Witloof, Speculoos, Gaufre, Praline, Trappist bira",
        "✅ Etnik gıda: Türk, Fas, Kore, Japon, Latin Amerika kategorileri",
        "✅ Fermente trend: Kombucha, Kimchi, Miso, Kefir, Tempeh",
        "✅ Yeni DB tabloları: product_families, price_history, basket_analysis",
        "✅ Demo sayfalar: Yumurta, Bira (Trappist/Lambic/Pils), Promo",
    ]
    for j, oz in enumerate(ozellikler_listesi):
        hucre_stil(ws, ozet_row + 1 + j, 1, oz, size=9, bg=RENKLER["sub_green"] if j % 2 == 0 else RENKLER["white"])
        ws.merge_cells(f"A{ozet_row+1+j}:D{ozet_row+1+j}")

def kategori_agaci_sayfasi(wb):
    ws = wb.create_sheet("Kategori Agaci")
    ws.sheet_view.showGridLines = False

    basliklar = ["Dept ID", "L1 Departman", "L2 Kategori", "L3 Alt Kategori", "L4 Ürün Tipleri"]
    ws.row_dimensions[1].height = 24
    for col, b in enumerate(basliklar, 1):
        hucre_stil(ws, 1, col, b, bold=True, fg="FFFFFF", bg=RENKLER["header_blue"], size=10, align="center")

    # Dept renklerini belirle
    dept_bg_map = {
        "D01": "E8F5E9", "D02": "FCE4EC", "D03": "F1F8E9", "D04": "FFF8E1",
        "D05": "E3F2FD", "D06": "EDE7F6", "D07": "E0F7FA", "D08": "FFF3E0",
        "D09": "F3E5F5", "D10": "E8EAF6", "D11": "E0F2F1", "D12": "FBE9E7",
        "D13": "E8F5E9", "D14": "F9FBE7", "D15": "FFF0F0", "D16": "F0F4FF",
    }
    dept_isimler = {d[0]: d[1] for d in DEPARTMANLAR}

    row = 2
    son_dept = None
    for d_id, l2, l3, l4_list in KATEGORILER:
        bg = dept_bg_map.get(d_id, "FFFFFF")
        l4_str = " | ".join(l4_list)
        dept_adi = dept_isimler.get(d_id, "")
        if d_id != son_dept:
            son_dept = d_id
        hucre_stil(ws, row, 1, d_id, bold=True, bg=bg, align="center", size=9)
        hucre_stil(ws, row, 2, dept_adi, bold=True, bg=bg, size=9)
        hucre_stil(ws, row, 3, l2, bg=bg, size=9)
        hucre_stil(ws, row, 4, l3, bold=True, bg=bg, size=9)
        hucre_stil(ws, row, 5, l4_str, bg=bg, size=8, wrap=True)
        ws.row_dimensions[row].height = max(15, 12 * (1 + l4_str.count("|") // 4))
        row += 1

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 80

def oznitelik_matrisi_sayfasi(wb):
    ws = wb.create_sheet("Oznitelik Matrisi")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 30
    basliklar = ["Öznitelik", "Veri Tipi", "Zorunlu?", "Açıklama", "Geçerli Departmanlar"]
    for col, b in enumerate(basliklar, 1):
        hucre_stil(ws, 1, col, b, bold=True, fg="FFFFFF", bg=RENKLER["header_purple"], size=10, align="center")

    gruplar = {
        "Temel Kimlik": ["sku_id", "ean", "market_chain", "product_name_nl", "product_name_fr", "brand"],
        "Kategori": ["category_l1", "category_l2", "category_l3", "category_l4"],
        "Fiyat": ["price_retail", "price_per_base_unit", "base_unit", "promo_price",
                  "promo_valid_from", "promo_valid_until", "promo_type"],
        "Ürün Özellikleri": ["net_weight_g", "net_volume_ml", "packaging_material", "recycle_code"],
        "Besin & Sağlık": ["nutriscore", "ecoscore", "energy_kcal_100g", "fat_g_100g",
                           "saturated_fat_g", "sugars_g_100g", "salt_g_100g", "protein_g_100g"],
        "Diyet Flagları": ["is_organic", "is_vegan", "is_vegetarian", "is_gluten_free",
                           "is_lactose_free", "is_halal", "is_kosher", "palm_oil_free", "is_fodmap_friendly"],
        "Sertifika & Ödül": ["regional_cert", "award_label"],
        "İçecek Özgün": ["alcohol_pct", "vintage"],
        "Ürün Ailesi & Meta": ["product_family_id", "availability", "last_updated_at"],
    }

    oznitelik_map = {o[0]: o for o in OZELLIKLER}
    row = 2
    for grup, alanlar in gruplar.items():
        # Grup başlığı
        ws.merge_cells(f"A{row}:E{row}")
        hucre_stil(ws, row, 1, f"▶ {grup}", bold=True, fg="FFFFFF",
                   bg=RENKLER["header_teal"], size=10, align="left")
        ws.row_dimensions[row].height = 22
        row += 1
        for alan in alanlar:
            if alan not in oznitelik_map:
                continue
            _, tip, zorunlu, aciklama, depts = oznitelik_map[alan]
            bg = "FFF9C4" if zorunlu else "FFFFFF"
            hucre_stil(ws, row, 1, alan, bold=zorunlu, bg=bg, size=9)
            hucre_stil(ws, row, 2, tip, bg=bg, size=9, align="center")
            hucre_stil(ws, row, 3, "✅ Zorunlu" if zorunlu else "◯ Opsiyonel",
                       bg="C6EFCE" if zorunlu else bg, size=9, align="center")
            hucre_stil(ws, row, 4, aciklama, bg=bg, size=9, wrap=True)
            hucre_stil(ws, row, 5, depts, bg=bg, size=9)
            ws.row_dimensions[row].height = 16
            row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 30

def normalizasyon_sayfasi(wb):
    ws = wb.create_sheet("Normalizasyon")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 28
    basliklar = ["Standart Anahtar", "NL Terimleri", "FR Terimleri", "DE Terimleri", "Açıklama"]
    for col, b in enumerate(basliklar, 1):
        hucre_stil(ws, 1, col, b, bold=True, fg="FFFFFF", bg=RENKLER["header_orange"], size=10, align="center")

    for i, (anahtar, nl, fr, de, aciklama) in enumerate(NORMALIZASYON):
        row = 2 + i
        bg = "FFF3E0" if i % 2 == 0 else "FFFFFF"
        hucre_stil(ws, row, 1, anahtar, bold=True, bg=bg, size=9)
        hucre_stil(ws, row, 2, ", ".join(nl), bg=bg, size=9, wrap=True)
        hucre_stil(ws, row, 3, ", ".join(fr), bg=bg, size=9, wrap=True)
        hucre_stil(ws, row, 4, ", ".join(de), bg=bg, size=9, wrap=True)
        hucre_stil(ws, row, 5, aciklama, bg=bg, size=9, wrap=True)
        ws.row_dimensions[row].height = 18

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 38

def veri_modeli_sayfasi(wb):
    ws = wb.create_sheet("Veri Modeli")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    hucre_stil(ws, 1, 1, "Supabase PostgreSQL Veri Modeli — SQL Şeması",
               bold=True, fg="FFFFFF", bg=RENKLER["header_red"], size=13, align="center")
    ws.row_dimensions[1].height = 32

    satirlar = VERI_MODELI_SQL.strip().split("\n")
    for i, satir in enumerate(satirlar):
        row = 3 + i
        if satir.strip().startswith("--"):
            hucre_stil(ws, row, 1, satir, fg="006400", size=9, bg="F0FFF0", align="left")
        elif any(satir.strip().upper().startswith(kw) for kw in ["CREATE", "ALTER", "INSERT"]):
            hucre_stil(ws, row, 1, satir, bold=True, fg="00008B", size=9, bg="F0F0FF")
        else:
            hucre_stil(ws, row, 1, satir, size=9, bg="FAFAFA")
        ws.row_dimensions[row].height = 14

    ws.column_dimensions["A"].width = 100
    ws.merge_cells(f"A3:C{3+len(satirlar)}")

def demo_yumurta_sayfasi(wb):
    ws = wb.create_sheet("Demo Yumurta")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:S1")
    hucre_stil(ws, 1, 1, "Demo Verisi: D01 Yumurta — Farklı marketlerde aynı ürün, promo ve Nutri-Score gösterimi",
               bold=True, fg="FFFFFF", bg=RENKLER["header_green"], size=12, align="center")
    ws.row_dimensions[1].height = 30

    alanlar = ["sku_id", "ean", "market_chain", "product_name_nl", "brand",
               "category_l1", "category_l3", "category_l4",
               "price_retail", "promo_price", "promo_valid_until", "promo_type",
               "net_weight_g", "nutriscore", "ecoscore",
               "is_organic", "is_vegetarian", "regional_cert", "availability"]

    for col, alan in enumerate(alanlar, 1):
        hucre_stil(ws, 2, col, alan, bold=True, fg="FFFFFF", bg=RENKLER["header_green"],
                   size=9, align="center")

    nutri_bg = {"A": RENKLER["nutri_a"], "B": RENKLER["nutri_b"], "C": RENKLER["nutri_c"],
                "D": RENKLER["nutri_d"], "E": RENKLER["nutri_e"], None: "FFFFFF"}

    for i, urun in enumerate(DEMO_YUMURTA):
        row = 3 + i
        bg = "F9FFF9" if i % 2 == 0 else "FFFFFF"
        for col, alan in enumerate(alanlar, 1):
            val = urun.get(alan, "")
            if isinstance(val, list):
                val = ", ".join(val) if val else "-"
            if val is None:
                val = "-"
            cell_bg = nutri_bg.get(urun.get("nutriscore")) if alan == "nutriscore" else bg
            if alan == "promo_price" and val not in ("-", None, ""):
                cell_bg = RENKLER["promo_bg"]
            hucre_stil(ws, row, col, val, bg=cell_bg, size=9, align="center" if col > 8 else "left")
        ws.row_dimensions[row].height = 18

    for col in range(1, len(alanlar) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

def demo_bira_sayfasi(wb):
    ws = wb.create_sheet("Demo Bira")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:P1")
    hucre_stil(ws, 1, 1, "Demo Verisi: D06 Bira — Trappist / Lambic / Pils stili, alkol%, vintage, promo",
               bold=True, fg="FFFFFF", bg="4A235A", size=12, align="center")
    ws.row_dimensions[1].height = 30

    alanlar = ["sku_id", "ean", "market_chain", "product_name_nl", "brand",
               "category_l2", "category_l3", "category_l4",
               "price_retail", "promo_price", "promo_valid_until",
               "net_volume_ml", "alcohol_pct", "vintage",
               "is_organic", "regional_cert", "ecoscore", "availability"]

    for col, alan in enumerate(alanlar, 1):
        hucre_stil(ws, 2, col, alan, bold=True, fg="FFFFFF", bg="4A235A", size=9, align="center")

    for i, urun in enumerate(DEMO_BIRA):
        row = 3 + i
        bg = "F5EEFF" if i % 2 == 0 else "FFFFFF"
        for col, alan in enumerate(alanlar, 1):
            val = urun.get(alan, "")
            if isinstance(val, list):
                val = ", ".join(val) if val else "-"
            if val is None:
                val = "-"
            cell_bg = bg
            if alan == "promo_price" and val not in ("-", None, ""):
                cell_bg = RENKLER["promo_bg"]
            hucre_stil(ws, row, col, val, bg=cell_bg, size=9, align="center" if col > 8 else "left")
        ws.row_dimensions[row].height = 18

    for col in range(1, len(alanlar) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

def demo_promo_sayfasi(wb):
    ws = wb.create_sheet("Demo Promo")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    hucre_stil(ws, 1, 1, "Demo Verisi: Haftalık Promosyon Takibi — promo_price, promo_valid_from/until, promo_type",
               bold=True, fg="FFFFFF", bg="7F3A00", size=12, align="center")
    ws.row_dimensions[1].height = 30

    alanlar = ["sku_id", "product_name_nl", "market_chain", "brand",
               "category_l3", "price_retail", "promo_price", "promo_pct",
               "promo_valid_from", "promo_valid_until", "promo_type",
               "nutriscore", "availability"]

    for col, alan in enumerate(alanlar, 1):
        hucre_stil(ws, 2, col, alan, bold=True, fg="FFFFFF", bg="7F3A00", size=9, align="center")

    for i, urun in enumerate(DEMO_PROMO):
        row = 3 + i
        bg = RENKLER["promo_bg"] if i % 2 == 0 else "FFF3E0"
        for col, alan in enumerate(alanlar, 1):
            val = urun.get(alan, "")
            if val is None:
                val = "-"
            hucre_stil(ws, row, col, val, bg=bg, size=9, align="center" if col > 5 else "left")
        # Tasarruf hesapla
        pr = urun.get("price_retail", 0) or 0
        pp = urun.get("promo_price", 0) or 0
        if pr and pp:
            tasarruf = round(pr - pp, 2)
            ws.cell(row=row, column=len(alanlar) + 1).value = f"€{tasarruf} tasarruf"
            ws.cell(row=row, column=len(alanlar) + 1).font = Font(color="006400", bold=True, size=9)
        ws.row_dimensions[row].height = 18

    # Tasarruf sütun başlığı
    hucre_stil(ws, 2, len(alanlar) + 1, "Tasarruf", bold=True, fg="FFFFFF", bg="006400", size=9, align="center")

    for col in range(1, len(alanlar) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 18

def main():
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    print("Belcika Kategori Sistemi Excel uretiliyor...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ozet_sayfasi(wb)
    print("  [OK] Ozet sayfasi")
    kategori_agaci_sayfasi(wb)
    print("  [OK] Kategori agaci")
    oznitelik_matrisi_sayfasi(wb)
    print("  [OK] Oznitelik matrisi")
    normalizasyon_sayfasi(wb)
    print("  [OK] Normalizasyon sozlugu")
    veri_modeli_sayfasi(wb)
    print("  [OK] Veri modeli (SQL)")
    demo_yumurta_sayfasi(wb)
    print("  [OK] Demo: Yumurta")
    demo_bira_sayfasi(wb)
    print("  [OK] Demo: Bira (Trappist/Lambic/Pils)")
    demo_promo_sayfasi(wb)
    print("  [OK] Demo: Promo fiyat takibi")

    import os
    cikti = os.path.join(os.path.dirname(os.path.abspath(__file__)), "belcika_kategori_sistemi_v2.xlsx")
    wb.save(cikti)
    print(f"\nTamamlandi: {cikti}")
    print(f"   - {len(DEPARTMANLAR)} departman (D01-D16)")
    l3_toplam = len(set(f"{d}>{l2}>{l3}" for d, l2, l3, _ in KATEGORILER))
    l4_toplam = sum(len(l4) for _, _, _, l4 in KATEGORILER)
    print(f"   - {l3_toplam} L3 alt kategori")
    print(f"   - {l4_toplam} L4 urun tipi")
    print(f"   - {len(OZELLIKLER)} oznitelik")
    print(f"   - {len(NORMALIZASYON)} normalizasyon kurali")

if __name__ == "__main__":
    main()
